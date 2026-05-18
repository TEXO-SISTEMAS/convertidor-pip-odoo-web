"""
Lógica de transformación de datos PIP a formato Odoo.
"""

import openpyxl
from datetime import datetime as _dt
from pathlib import Path
from difflib import SequenceMatcher

from utils import (
    convertir_fecha_pip_a_odoo,
    extraer_partes_numero_factura,
    calcular_terminos_pago,
    mapear_producto_odoo,
)

# Nombre exacto de la hoja en el Excel PIP
HOJA_PIP = "Resumen Detallado"

# Columnas esperadas en el Excel PIP (orden y nombres exactos)
COLUMNAS_PIP = [
    "N° de Documento",
    "CDC",
    "Tipo de Documento",
    "Fecha de Emisión",
    "Fecha de Vencimiento",
    "Estado",
    "RUC/CI",
    "Cliente",
    "Código de Producto",
    "Nombre del Producto",
    "Volumen",
    "Moneda",
    "Precio Unitario",
    "Descuento Unitario",
    "Exento",
    "Gravado 5%",
    "Gravado 10%",
    "TOTAL",
    "Pago",
    "Observación",
    "Fecha de creación",
]

# Nuevas columnas (18 en total)
COLUMNAS_ODOO = [
    "Contacto",
    "Fecha de factura",
    "Tipo comprobante",
    "Tipo Factura",
    "Talonario",
    "Timbrado",
    "Vencimiento del timbrado",
    "Suc",
    "Sec",
    "Nro",
    "Nro. de factura",
    "Términos de pago",
    "Diario",
    "Líneas de factura/Producto",
    "Líneas de factura/Etiqueta",
    "Líneas de factura/Cantidad",
    "Líneas de factura/Unidad de medida",
    "Líneas de factura/Precio unitario",
]

# Valores fijos
TIPO_COMPROBANTE = "Factura"
TALONARIO = "FACTURA PREIMPRESA"
TIMBRADO = "17087942"
VENCIMIENTO_TIMBRADO = "2050-03-13"
DIARIO = "Facturas de cliente"
UNIDAD_MEDIDA = "Unidad"


class ConvertidorPipOdoo:
    """
    Convierte un archivo Excel exportado del sistema PIP
    al formato de importación masiva de Odoo.
    """

    def __init__(self):
        self._libro_pip = None       # Workbook de entrada (openpyxl)
        self._hoja_pip = None        # Hoja "Resumen Detallado"
        self._indices = {}           # Mapa columna → índice (0-based) en el PIP
        self._filas_convertidas = [] # Lista de dicts con los datos transformados
        self._alertas = []           # Alertas no bloqueantes generadas durante la conversión
        self._nombres_odoo = []      # Lista de nombres correctos de Odoo
        self._contactos_no_mapeados = {}  # Dict: nombre_pip -> mejor_score
        self._cargar_nombres_odoo() # Cargar nombres correctos al iniciar

    # ------------------------------------------------------------------
    # 0. CARGA DE NOMBRES CORRECTOS DE ODOO
    # ------------------------------------------------------------------

    def _cargar_nombres_odoo(self):
        """
        Carga el archivo nombres_correctos.xlsx con los nombres correctos de Odoo.
        El archivo debe estar en la misma carpeta que el script o el .exe.
        """
        try:
            # Buscar el archivo en varias ubicaciones posibles
            posibles_rutas = [
                Path("nombres_correctos.xlsx"),  # Carpeta actual
                Path(__file__).parent / "nombres_correctos.xlsx",  # Carpeta del script
                Path(__file__).parent.parent / "nombres_correctos.xlsx",  # Carpeta padre
            ]
            
            archivo_nombres = None
            for ruta in posibles_rutas:
                if ruta.exists():
                    archivo_nombres = ruta
                    break
            
            if not archivo_nombres:
                print("[WARNING] No se encontró nombres_correctos.xlsx. Los nombres no se mapearán.")
                return
            
            # Cargar archivo
            wb = openpyxl.load_workbook(archivo_nombres, read_only=True, data_only=True)
            sheet = wb.active
            
            # Leer nombres (columna "Contacto", saltear encabezado)
            self._nombres_odoo = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Si tiene valor en la primera columna
                    nombre = str(row[0]).strip()
                    if nombre:
                        self._nombres_odoo.append(nombre)
            
            wb.close()
            
            print(f"[INFO] Cargados {len(self._nombres_odoo)} nombres de Odoo para mapeo.")
            
        except Exception as e:
            print(f"[WARNING] Error al cargar nombres_correctos.xlsx: {e}")
            print("[WARNING] Los nombres no se mapearán automáticamente.")

    def _buscar_nombre_similar(self, nombre_pip: str) -> str:
        """
        Busca el nombre más similar en la lista de nombres de Odoo.
        
        Args:
            nombre_pip: Nombre del cliente como aparece en PIP
            
        Returns:
            Nombre correcto de Odoo si se encuentra coincidencia > 60%,
            caso contrario retorna el nombre original del PIP limpio.
        """
        if not self._nombres_odoo:
            # No hay lista de nombres, retornar limpio
            return nombre_pip.strip()
        
        # Normalizar nombre PIP para búsqueda
        nombre_limpio = nombre_pip.strip().upper()
        
        # Caso 1: Coincidencia exacta (rápido)
        for nombre_odoo in self._nombres_odoo:
            if nombre_odoo.upper() == nombre_limpio:
                return nombre_odoo
        
        # Caso 2: Búsqueda por similitud
        mejor_coincidencia = None
        mejor_score = 0.0
        
        for nombre_odoo in self._nombres_odoo:
            # Calcular similitud usando SequenceMatcher
            score = SequenceMatcher(None, nombre_limpio, nombre_odoo.upper()).ratio()
            
            if score > mejor_score:
                mejor_score = score
                mejor_coincidencia = nombre_odoo
        
        # Si la similitud es > 60%, usar nombre de Odoo (bajado de 70% a 60%)
        if mejor_score >= 0.60:
            if mejor_score < 1.0:  # No es coincidencia exacta
                print(f"[INFO] Mapeado: '{nombre_pip}' → '{mejor_coincidencia}' (similitud: {mejor_score:.0%})")
            return mejor_coincidencia
        else:
            # No hay buena coincidencia, registrar para reporte
            if nombre_limpio not in self._contactos_no_mapeados:
                self._contactos_no_mapeados[nombre_limpio] = mejor_score
            print(f"[WARNING] No se encontró coincidencia para '{nombre_pip}' (mejor: {mejor_score:.0%})")
            return nombre_limpio

    # ------------------------------------------------------------------
    # 1. CARGA DEL EXCEL PIP
    # ------------------------------------------------------------------

    def cargar_excel_pip(self, ruta_archivo: str) -> tuple[bool, str]:
        """
        Abre y valida superficialmente el Excel PIP.

        Verifica:
          - Que el archivo sea .xlsx
          - Usa la primera hoja disponible (sin importar el nombre)
          - Que tenga las 21 columnas esperadas con nombres exactos

        Args:
            ruta_archivo: Ruta completa al archivo .xlsx del sistema PIP.

        Returns:
            (True, mensaje_ok) si el archivo es válido.
            (False, mensaje_error) si hay algún problema.
        """
        print(f"[INFO] Cargando archivo: {ruta_archivo}")

        # Validar extensión
        if not ruta_archivo.lower().endswith(".xlsx"):
            return False, "Solo se permiten archivos Excel (.xlsx)."

        # Abrir el libro
        try:
            self._libro_pip = openpyxl.load_workbook(ruta_archivo, read_only=True, data_only=True)
        except Exception as e:
            return False, f"No se pudo abrir el archivo: {e}"

        # Usar la primera hoja disponible (sin importar el nombre)
        self._hoja_pip = self._libro_pip.active
        nombre_hoja = self._hoja_pip.title
        print(f"[INFO] Usando hoja: '{nombre_hoja}'")

        # Leer encabezados de la primera fila y construir mapa columna → índice
        encabezados = [str(celda.value).strip() if celda.value is not None else "" for celda in next(self._hoja_pip.iter_rows(min_row=1, max_row=1))]

        print(f"[INFO] Columnas encontradas en PIP: {encabezados}")

        # Verificar que todas las columnas requeridas estén presentes
        columnas_faltantes = [col for col in COLUMNAS_PIP if col not in encabezados]
        if columnas_faltantes:
            return False, f"Faltan columnas en el archivo: {', '.join(columnas_faltantes)}"

        # Guardar índice de cada columna para acceso rápido por nombre
        self._indices = {col: encabezados.index(col) for col in COLUMNAS_PIP}

        # Contar filas de datos (excluyendo encabezado)
        total_filas = sum(1 for _ in self._hoja_pip.iter_rows(min_row=2))
        print(f"[INFO] Archivo cargado OK. Filas de datos detectadas: {total_filas}")

        return True, f"Archivo cargado correctamente. {total_filas} facturas detectadas."

    # ------------------------------------------------------------------
    # 2. CONVERSIÓN FILA POR FILA
    # ------------------------------------------------------------------

    def convertir(self, tipo_impuesto: str = "10%", tipo_documento: str = "factura") -> tuple[bool, str]:
        """
        Procesa cada fila del Excel PIP y aplica las reglas de transformación.
        Consolida líneas con el mismo número de factura y código de producto.

        Debe llamarse después de cargar_excel_pip().

        Returns:
            (True, resumen) si la conversión fue exitosa.
            (False, mensaje_error) si ocurrió un error fatal.
        """
        if self._hoja_pip is None:
            return False, "No hay archivo cargado. Llamar primero a cargar_excel_pip()."

        self._alertas = []
        errores = []
        filas_temporales = []

        print("[INFO] Iniciando conversión...")

        # PASO 1: Convertir todas las filas
        for num_fila, fila in enumerate(self._hoja_pip.iter_rows(min_row=2, values_only=True), start=2):
            # Saltear filas completamente vacías
            if all(celda is None for celda in fila):
                print(f"[DEBUG] Fila {num_fila}: vacía, se omite.")
                continue

            try:
                filas_odoo = self._transformar_fila(fila, num_fila, tipo_impuesto, tipo_documento)
                filas_temporales.extend(filas_odoo)  # Agregar todas las líneas
            except UnicodeEncodeError:
                # Error de encoding en print de debug: no es un error de datos, ignorar
                pass
            except ValueError as e:
                # Error de dato inválido: se registra pero continúa con las demás filas
                msg = f"Fila {num_fila}: {e}"
                errores.append(msg)
            print(f"[DEBUG] Fila {num_fila}: procesada")

        # PASO 2: Consolidar líneas por factura y producto
        print("[INFO] Consolidando líneas de productos...")
        self._filas_convertidas = self._consolidar_lineas(filas_temporales)
        
        lineas_originales = len(filas_temporales)
        lineas_consolidadas = len(self._filas_convertidas)
        if lineas_consolidadas < lineas_originales:
            print(f"[INFO] Consolidación: {lineas_originales} líneas → {lineas_consolidadas} facturas")

        # Cerrar el workbook de entrada: libera el lock del archivo en Windows
        try:
            self._libro_pip.close()
        except Exception:
            pass
        self._libro_pip = None
        self._hoja_pip = None

        if errores:
            detalle = "\n".join(errores)
            return False, f"Se encontraron errores en la conversión:\n{detalle}"

        resumen = f"Conversión exitosa. {len(self._filas_convertidas)} facturas procesadas."
        if self._alertas:
            resumen += f" ({len(self._alertas)} alertas, ver log)."
        print(f"[INFO] {resumen}")
        
        # Generar reporte de contactos no mapeados
        if self._contactos_no_mapeados:
            self._generar_reporte_no_mapeados()
        
        return True, resumen

    def _generar_reporte_no_mapeados(self):
        """
        Genera un archivo de texto con la lista de contactos que no se pudieron mapear.
        """
        try:
            archivo_reporte = "contactos_no_mapeados.txt"
            
            with open(archivo_reporte, 'w', encoding='utf-8') as f:
                f.write("="*80 + "\n")
                f.write("CONTACTOS NO MAPEADOS - CONVERTIDOR PIP A ODOO\n")
                f.write("="*80 + "\n\n")
                f.write(f"Total de contactos sin mapear: {len(self._contactos_no_mapeados)}\n")
                f.write(f"Umbral de similitud: 60%\n\n")
                f.write("Estos contactos no alcanzaron el 60% de similitud con ningún nombre\n")
                f.write("de la lista nombres_correctos.xlsx.\n\n")
                f.write("RECOMENDACIÓN:\n")
                f.write("1. Buscar los nombres correctos en Odoo\n")
                f.write("2. Agregarlos al archivo nombres_correctos.xlsx\n")
                f.write("3. Volver a ejecutar el convertidor\n\n")
                f.write("="*80 + "\n\n")
                
                # Ordenar por similitud (de menor a mayor)
                contactos_ordenados = sorted(
                    self._contactos_no_mapeados.items(),
                    key=lambda x: x[1]
                )
                
                for i, (nombre, score) in enumerate(contactos_ordenados, 1):
                    f.write(f"{i:3}. {nombre:<50} (similitud máxima: {score:.0%})\n")
            
            print(f"[INFO] Generado reporte: {archivo_reporte}")
            
        except Exception as e:
            print(f"[WARNING] No se pudo generar reporte de contactos no mapeados: {e}")

    def _consolidar_lineas(self, filas: list[dict]) -> list[dict]:
        """
        Consolida líneas con el mismo número de factura y código de producto.
        NO consolida líneas marcadas con _no_consolidar (caso Gravado 5% + Exento).
        NO consolida líneas de rec01 (Recupero) ya que pueden ser conceptos diferentes.
        
        Args:
            filas: Lista de diccionarios con las filas convertidas
            
        Returns:
            Lista de diccionarios consolidados
        """
        from collections import defaultdict
        
        # Separar líneas que NO deben consolidarse
        lineas_no_consolidar = []
        lineas_por_consolidar = []
        
        for fila in filas:
            # No consolidar si:
            # 1. Tiene flag _no_consolidar (Gravado 5% + Exento)
            # 2. Es producto rec01 (Recupero - diferentes conceptos)
            if fila.get("_no_consolidar", False):
                # Preservar _linea_numero, quitar solo _no_consolidar
                fila_limpia = {k: v for k, v in fila.items() if k != "_no_consolidar"}
                lineas_no_consolidar.append(fila_limpia)
            elif fila.get("Líneas de factura/Producto") == "rec01":
                # rec01 NO se consolida (Limpieza, GC, Consumo, etc. son diferentes)
                lineas_no_consolidar.append(fila)
            else:
                lineas_por_consolidar.append(fila)
        
        # Agrupar por (Nro. factura, Código Producto) solo las consolidables
        grupos = defaultdict(list)
        
        for fila in lineas_por_consolidar:
            clave = (
                fila["Nro. de factura"],
                fila["Líneas de factura/Producto"]
            )
            grupos[clave].append(fila)
        
        # Consolidar cada grupo
        filas_consolidadas = []
        
        for (nro_factura, cod_producto), lineas in grupos.items():
            if len(lineas) == 1:
                # Una sola línea, no hay nada que consolidar
                filas_consolidadas.append(lineas[0])
            else:
                # Múltiples líneas del mismo producto en la misma factura
                # Consolidar sumando cantidades y precios
                fila_consolidada = lineas[0].copy()
                
                cantidad_total = sum(
                    linea["Líneas de factura/Cantidad"] 
                    for linea in lineas
                )
                
                precio_total = sum(
                    linea["Líneas de factura/Precio unitario"] 
                    for linea in lineas
                )
                
                fila_consolidada["Líneas de factura/Cantidad"] = cantidad_total
                fila_consolidada["Líneas de factura/Precio unitario"] = precio_total
                
                print(f"[INFO] Consolidado: Factura {nro_factura} | {cod_producto} | {len(lineas)} líneas → Cant: {cantidad_total}, Precio: {precio_total}")
                
                filas_consolidadas.append(fila_consolidada)
        
        # Combinar respetando el orden original de aparición
        # Reconstruir en el orden en que aparecieron en el archivo PIP
        resultado = []
        consolidadas_usadas = set()
        no_consolidar_idx = 0
        
        for fila in filas:
            nro = fila.get("Nro. de factura", "")
            prod = fila.get("Líneas de factura/Producto", "")
            clave = (nro, prod)
            
            if fila.get("_no_consolidar", False) or prod == "rec01":
                # Buscar la versión limpia en lineas_no_consolidar
                for f in lineas_no_consolidar:
                    if f.get("Nro. de factura") == nro and f.get("Líneas de factura/Producto") == prod and f not in resultado:
                        resultado.append(f)
                        break
            else:
                # Buscar la versión consolidada
                for f in filas_consolidadas:
                    if f.get("Nro. de factura") == nro and f.get("Líneas de factura/Producto") == prod and f not in resultado:
                        resultado.append(f)
                        break
        
        return resultado

    def _mapear_tipo_documento(self, tipo_doc_pip: str) -> tuple[str, str, str, str]:
        """
        Mapea el tipo de documento del PIP al tipo de comprobante, talonario,
        timbrado y vencimiento del timbrado de Odoo.

        Args:
            tipo_doc_pip: Tipo de documento como aparece en PIP

        Returns:
            Tupla (tipo_comprobante, talonario, timbrado, vencimiento_timbrado)
        """
        if not tipo_doc_pip:
            return ("Factura", "FACTURA PREIMPRESA", "17087942", "2050-03-13")

        tipo_normalizado = str(tipo_doc_pip).strip().lower()

        if "nota de crédito" in tipo_normalizado or "nota de credito" in tipo_normalizado:
            return ("Nota de Crédito", "NOTA DE CRÉDITO PREIMPRESO", "17087942", "2037-12-31")
        elif "factura" in tipo_normalizado:
            return ("Factura", "FACTURA PREIMPRESA", "17087942", "2050-03-13")
        else:
            print(f"[WARNING] Tipo de documento desconocido: '{tipo_doc_pip}' - usando 'Factura electrónica'")
            return ("Factura", "FACTURA PREIMPRESA", "17087942", "2050-03-13")

    def _transformar_fila(self, fila: tuple, num_fila: int, tipo_impuesto: str = "10%", tipo_documento: str = "factura") -> dict:
        """
        Aplica todas las reglas de transformación a una sola fila del PIP.
        Genera un dict con las 18 columnas requeridas por Odoo.
        """
        def col(nombre):
            """Obtiene el valor de una columna PIP por nombre."""
            return fila[self._indices[nombre]]

        # Omitir filas anuladas
        estado = col("Estado")
        if estado and str(estado).strip().lower() == "anulado":
            print(f"[INFO] Fila {num_fila}: Estado=Anulado → omitida")
            return []

        # --- Campos del PIP ---
        numero_doc = col("N° de Documento")
        tipo_doc_pip = col("Tipo de Documento")  # NUEVO: Leer tipo de documento
        cliente = col("Cliente")
        fecha_emision_raw = col("Fecha de Emisión")
        fecha_vencimiento_raw = col("Fecha de Vencimiento")
        codigo_producto = col("Código de Producto")
        nombre_producto = col("Nombre del Producto")
        volumen = col("Volumen")
        precio_unitario = col("Precio Unitario")
        exento = col("Exento")  # NUEVO: Leer exento
        gravado_5 = col("Gravado 5%")  # NUEVO: Leer gravado 5%
        gravado_10 = col("Gravado 10%")  # NUEVO: Leer gravado 10%

        # Validar campos obligatorios
        if not numero_doc:
            raise ValueError("N° de Documento no puede estar vacío.")
        if not cliente:
            raise ValueError("Cliente no puede estar vacío.")
        if not fecha_emision_raw:
            raise ValueError("Fecha de Emisión no puede estar vacía.")

        # Normalizar fechas
        fecha_emision_str = _normalizar_fecha(fecha_emision_raw)
        fecha_vencimiento_str = _normalizar_fecha(fecha_vencimiento_raw)

        # Extraer partes del número de factura
        suc, sec, nro = extraer_partes_numero_factura(str(numero_doc))

        # Determinar tipo de factura (Crédito/Contado)
        tipo_factura = "Contado" if (not fecha_vencimiento_str or fecha_vencimiento_str == "-") else "Credito"

        # Calcular términos de pago
        terminos_pago = calcular_terminos_pago(fecha_emision_str, fecha_vencimiento_str)

        # Mapear producto de PIP a Odoo
        codigo_odoo, etiqueta_odoo = mapear_producto_odoo(
            str(codigo_producto) if codigo_producto else "",
            str(nombre_producto) if nombre_producto else ""
        )
        
        if tipo_documento == "nota_credito":
            tipo_comprobante = "Nota de Crédito"
            talonario = "NOTA DE CRÉDITO PREIMPRESO"
            timbrado = "17087942"
            vencimiento_timbrado = "2037-12-31"
        else:
            tipo_comprobante = "Factura"
            talonario = "FACTURA PREIMPRESA"
            timbrado = "17087942"
            vencimiento_timbrado = "2050-03-13"

        # Construir fila base Odoo (18 columnas)
        fila_base = {
            "Contacto": self._buscar_nombre_similar(str(cliente)),  # Mapear nombre
            "Fecha de factura": convertir_fecha_pip_a_odoo(fecha_emision_str),
            "Tipo comprobante": tipo_comprobante,
            "Tipo Factura": tipo_factura,
            "Talonario": talonario,
            "Timbrado": timbrado,
            "Vencimiento del timbrado": vencimiento_timbrado,
            "Suc": suc,
            "Sec": sec,
            "Nro": nro,
            "Nro. de factura": str(numero_doc).strip(),
            "Términos de pago": terminos_pago,
            "Diario": DIARIO,
            "Líneas de factura/Producto": codigo_odoo,
            "Líneas de factura/Etiqueta": etiqueta_odoo,
            "Líneas de factura/Cantidad": volumen if volumen else 1,
            "Líneas de factura/Unidad de medida": UNIDAD_MEDIDA,
            "Líneas de factura/Precio unitario": precio_unitario if precio_unitario else 0,
        }
        
        # Debug: mostrar valores de impuesto
        if exento or gravado_5 or gravado_10:
            print(f"[DEBUG] Fila {num_fila}: Exento={exento}, Gravado5%={gravado_5}, Gravado10%={gravado_10}")

        if tipo_impuesto == "5%":
            print(f"[INFO] Fila {num_fila}: tipo_impuesto=5% → generando 2 líneas (Gravado 5% + Exenta)")
            nombre_pip_str = str(nombre_producto) if nombre_producto else ""

            _mapa_exento = {
                "flo1":   "FLO1EX",
                "flo2":   "flo2EXE",
                "flo3":   "flo3exe",
                "flo4":   "flo4exe",
                "flo5":   "flo5exe",
                "flo6":   "flo6exe",
                "venter": "venterexe",
            }
            codigo_odoo_exento = _mapa_exento.get(codigo_odoo.lower(), codigo_odoo)

            linea1 = fila_base.copy()
            linea1["Líneas de factura/Etiqueta"] = nombre_pip_str
            linea1["Líneas de factura/Cantidad"] = 1
            linea1["Líneas de factura/Precio unitario"] = gravado_5 or 0
            linea1["_no_consolidar"] = True
            linea1["_linea_numero"] = 1

            linea2 = fila_base.copy()
            linea2["Líneas de factura/Producto"] = codigo_odoo_exento
            linea2["Líneas de factura/Etiqueta"] = nombre_pip_str.strip()
            linea2["Líneas de factura/Cantidad"] = 1
            linea2["Líneas de factura/Precio unitario"] = exento or 0
            linea2["_no_consolidar"] = True
            linea2["_linea_numero"] = 2

            return [linea1, linea2]
        else:
            # tipo_impuesto == "10%": 1 sola línea normal
            return [fila_base]

    # ------------------------------------------------------------------
    # 3. GUARDAR EXCEL ODOO
    # ------------------------------------------------------------------

    def guardar_excel_odoo(self, ruta_salida: str) -> tuple[bool, str]:
        """
        Genera el archivo Excel de salida con el formato requerido por Odoo.

        La hoja se llama "Sheet1" y tiene los encabezados en la primera fila
        seguidos de los datos convertidos.

        Args:
            ruta_salida: Ruta completa donde guardar el archivo .xlsx.

        Returns:
            (True, mensaje_ok) si se guardó correctamente.
            (False, mensaje_error) si ocurrió un error.
        """
        if not self._filas_convertidas:
            return False, "No hay datos convertidos. Ejecutar convertir() primero."

        print(f"[INFO] Guardando archivo de salida: {ruta_salida}")

        libro_salida = openpyxl.Workbook()
        hoja = libro_salida.active
        hoja.title = "Sheet1"

        # Primera fila: encabezados en el orden exacto requerido por Odoo
        hoja.append(COLUMNAS_ODOO)

        # Filas de datos: escribir según el tipo de línea
        facturas_escritas = set()  # Registrar facturas ya escritas
        
        for num_row, fila_dict in enumerate(self._filas_convertidas, start=2):
            linea_numero = fila_dict.get("_linea_numero", 1)
            nro_factura = fila_dict.get("Nro. de factura", "")
            
            # Determinar si esta fila es secundaria:
            # - Líneas 2 y 3 del modo 5% (tienen _linea_numero)
            # - Cualquier línea cuyo número de factura ya fue escrito (modo 10% con rec01)
            es_linea_secundaria = (linea_numero in [2, 3]) or (nro_factura in facturas_escritas)
            
            if not es_linea_secundaria:
                facturas_escritas.add(nro_factura)
            
            for col_idx, col_nombre in enumerate(COLUMNAS_ODOO, start=1):
                if es_linea_secundaria:
                    # Columnas 1-13: vacías
                    if col_idx <= 13:
                        hoja.cell(row=num_row, column=col_idx, value=None)
                    # Columnas 14-18: datos del producto
                    else:
                        valor = fila_dict.get(col_nombre)
                        hoja.cell(row=num_row, column=col_idx, value=valor)
                else:
                    # Primera línea de la factura: todos los datos
                    valor = fila_dict.get(col_nombre)
                    hoja.cell(row=num_row, column=col_idx, value=valor)

        try:
            libro_salida.save(ruta_salida)
        except Exception as e:
            return False, f"No se pudo guardar el archivo: {e}"

        print(f"[INFO] Archivo guardado OK: {ruta_salida}")
        return True, f"Archivo guardado correctamente en:\n{ruta_salida}"

    # ------------------------------------------------------------------
    # Accesores
    # ------------------------------------------------------------------

    @property
    def alertas(self) -> list[str]:
        """Lista de alertas no bloqueantes generadas durante la conversión."""
        return list(self._alertas)

    @property
    def total_convertidas(self) -> int:
        """Cantidad de filas convertidas exitosamente."""
        return len(self._filas_convertidas)


# ------------------------------------------------------------------
# Helpers de módulo
# ------------------------------------------------------------------

def _normalizar_fecha(valor) -> str:
    """
    Convierte un valor de celda a string con formato DD/MM/YYYY.

    openpyxl puede devolver la fecha como:
      - str  "30/01/2026"  → ya está en formato correcto
      - datetime           → se formatea a DD/MM/YYYY
      - None o "-"         → se devuelve "-"
    """
    if valor is None:
        return "-"

    if isinstance(valor, _dt):
        return valor.strftime("%d/%m/%Y")

    return str(valor).strip()
