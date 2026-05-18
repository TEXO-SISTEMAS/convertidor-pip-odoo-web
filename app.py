import sys
import os
import json
import uuid
import tempfile

# Forzar UTF-8 en stdout para evitar errores con caracteres especiales en Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

from pathlib import Path
from flask import Flask, request, render_template, send_file, flash, redirect, url_for
import openpyxl

# Agregar src/ al path
sys.path.insert(0, str(Path(__file__).parent / "src"))

from converter import ConvertidorPipOdoo
from utils import mapear_producto_odoo

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "cambiar-en-produccion")

EXTENSIONES_PERMITIDAS = {"xlsx"}
MAPPINGS_PATH = Path(__file__).parent / "mappings.json"

# Columnas del Excel PIP que necesitamos para el análisis
COL_CODIGO = "Código de Producto"
COL_NOMBRE = "Nombre del Producto"
COL_ESTADO = "Estado"


def extension_valida(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in EXTENSIONES_PERMITIDAS


def _leer_mappings() -> list[dict]:
    if not MAPPINGS_PATH.exists():
        return []
    with open(MAPPINGS_PATH, encoding="utf-8") as f:
        return json.load(f)


def _guardar_mappings(mappings: list[dict]):
    with open(MAPPINGS_PATH, "w", encoding="utf-8") as f:
        json.dump(mappings, f, ensure_ascii=False, indent=2)


def _extraer_productos_no_mapeados(ruta_archivo: str) -> tuple[list[dict], str | None]:
    """
    Lee el Excel PIP y devuelve los productos únicos que no tienen mapeo conocido.
    Retorna (lista_no_mapeados, error_message).
    error_message es None si todo fue bien, o un string descriptivo si hubo problema.
    """
    try:
        wb = openpyxl.load_workbook(ruta_archivo, read_only=True, data_only=True)
        hoja = wb.active
        encabezados = [str(c.value).strip() if c.value else "" for c in next(hoja.iter_rows(min_row=1, max_row=1))]

        if COL_CODIGO not in encabezados or COL_NOMBRE not in encabezados:
            wb.close()
            return [], (
                "El archivo no parece ser un Excel exportado del PIP. "
                f"Se esperaban las columnas '{COL_CODIGO}' y '{COL_NOMBRE}'. "
                "Asegurate de subir el archivo original del PIP, no el convertido para Odoo."
            )

        idx_codigo = encabezados.index(COL_CODIGO)
        idx_nombre = encabezados.index(COL_NOMBRE)
        idx_estado = encabezados.index(COL_ESTADO) if COL_ESTADO in encabezados else None

        vistos = set()
        no_mapeados = []

        for fila in hoja.iter_rows(min_row=2, values_only=True):
            if all(c is None for c in fila):
                continue
            if idx_estado is not None:
                estado = str(fila[idx_estado] or "").strip().lower()
                if estado == "anulado":
                    continue

            codigo = str(fila[idx_codigo] or "").strip()
            nombre = str(fila[idx_nombre] or "").strip()

            if not codigo and not nombre:
                continue

            clave = (codigo.upper(), nombre.upper())
            if clave in vistos:
                continue
            vistos.add(clave)

            cod_odoo, nom_odoo = mapear_producto_odoo(codigo, nombre)

            # Si el resultado es igual al input, no matcheó ninguna regla
            if cod_odoo == codigo and nom_odoo == nombre:
                no_mapeados.append({"codigo_pip": codigo, "nombre_pip": nombre})

        wb.close()
        return no_mapeados, None

    except Exception as e:
        return [], f"Error al leer el archivo: {e}"


# ------------------------------------------------------------------
# Rutas principales
# ------------------------------------------------------------------

@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")


@app.route("/convertir", methods=["POST"])
def convertir():
    archivo = request.files.get("archivo_pip")
    tipo_impuesto = request.form.get("tipo_impuesto", "10%")
    tipo_documento = request.form.get("tipo_documento", "factura")

    if not archivo or archivo.filename == "":
        flash("Seleccioná un archivo Excel (.xlsx) para convertir.")
        return redirect(url_for("index"))

    if not extension_valida(archivo.filename):
        flash("Solo se permiten archivos Excel (.xlsx).")
        return redirect(url_for("index"))

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_entrada:
        archivo.save(tmp_entrada.name)
        ruta_entrada = tmp_entrada.name

    ruta_salida = ruta_entrada.replace(".xlsx", "_odoo.xlsx")

    try:
        convertidor = ConvertidorPipOdoo()

        ok, msg = convertidor.cargar_excel_pip(ruta_entrada)
        if not ok:
            flash(f"Error al cargar el archivo: {msg}")
            return redirect(url_for("index"))

        ok, msg = convertidor.convertir(tipo_impuesto=tipo_impuesto, tipo_documento=tipo_documento)
        if not ok:
            flash(f"Error en la conversión: {msg}")
            return redirect(url_for("index"))

        ok, msg = convertidor.guardar_excel_odoo(ruta_salida)
        if not ok:
            flash(f"Error al guardar el resultado: {msg}")
            return redirect(url_for("index"))

        nombre_descarga = f"odoo_{Path(archivo.filename).stem}.xlsx"
        return send_file(
            ruta_salida,
            as_attachment=True,
            download_name=nombre_descarga,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    finally:
        try:
            os.unlink(ruta_entrada)
        except Exception:
            pass


@app.route("/analizar", methods=["POST"])
def analizar():
    """
    Lee el archivo PIP y muestra qué productos no tienen mapeo,
    permitiendo agregarlos antes de convertir.
    """
    archivo = request.files.get("archivo_pip")
    tipo_impuesto = request.form.get("tipo_impuesto", "10%")
    tipo_documento = request.form.get("tipo_documento", "factura")

    if not archivo or archivo.filename == "":
        flash("Seleccioná un archivo Excel (.xlsx).")
        return redirect(url_for("index"))

    if not extension_valida(archivo.filename):
        flash("Solo se permiten archivos Excel (.xlsx).")
        return redirect(url_for("index"))

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        archivo.save(tmp.name)
        ruta_tmp = tmp.name

    no_mapeados, error_formato = _extraer_productos_no_mapeados(ruta_tmp)

    try:
        os.unlink(ruta_tmp)
    except Exception:
        pass

    return render_template(
        "analizar.html",
        no_mapeados=no_mapeados,
        error_formato=error_formato,
        tipo_impuesto=tipo_impuesto,
        tipo_documento=tipo_documento,
        nombre_archivo=archivo.filename,
    )


# ------------------------------------------------------------------
# Rutas de gestión de productos
# ------------------------------------------------------------------

@app.route("/productos", methods=["GET"])
def productos():
    mappings = _leer_mappings()
    return render_template("productos.html", mappings=mappings)


def _detectar_tipo_match(valor: str) -> str:
    if valor.endswith("-"):
        return "codigo_prefijo"
    return "codigo"


@app.route("/productos/agregar", methods=["POST"])
def agregar_producto():
    valor_match = request.form.get("valor_match", "").strip().upper()
    codigo_odoo = request.form.get("codigo_odoo", "").strip()
    nombre_odoo = request.form.get("nombre_odoo", "").strip()

    if not valor_match or not codigo_odoo or not nombre_odoo:
        mappings = _leer_mappings()
        return render_template("productos.html", mappings=mappings, error="Todos los campos son obligatorios.")

    tipo_match = _detectar_tipo_match(valor_match)
    mappings = _leer_mappings()
    mappings.append({
        "id": str(uuid.uuid4()),
        "tipo_match": tipo_match,
        "valor_match": valor_match,
        "codigo_odoo": codigo_odoo,
        "nombre_odoo": nombre_odoo,
    })
    _guardar_mappings(mappings)
    return render_template("productos.html", mappings=mappings, exito=f"Producto '{nombre_odoo}' agregado correctamente.")


@app.route("/productos/editar/<mapping_id>", methods=["POST"])
def editar_producto(mapping_id):
    valor_match = request.form.get("valor_match", "").strip().upper()
    codigo_odoo = request.form.get("codigo_odoo", "").strip()
    nombre_odoo = request.form.get("nombre_odoo", "").strip()

    if not valor_match or not codigo_odoo or not nombre_odoo:
        mappings = _leer_mappings()
        return render_template("productos.html", mappings=mappings, error="Todos los campos son obligatorios.")

    tipo_match = _detectar_tipo_match(valor_match)
    mappings = _leer_mappings()
    for m in mappings:
        if m.get("id") == mapping_id:
            m["tipo_match"] = tipo_match
            m["valor_match"] = valor_match
            m["codigo_odoo"] = codigo_odoo
            m["nombre_odoo"] = nombre_odoo
            break
    _guardar_mappings(mappings)
    return render_template("productos.html", mappings=mappings, exito=f"Producto '{nombre_odoo}' actualizado correctamente.")


@app.route("/productos/eliminar/<mapping_id>", methods=["POST"])
def eliminar_producto(mapping_id):
    mappings = _leer_mappings()
    mappings = [m for m in mappings if m.get("id") != mapping_id]
    _guardar_mappings(mappings)
    return render_template("productos.html", mappings=mappings, exito="Producto eliminado.")


if __name__ == "__main__":
    app.run(debug=True, use_reloader=False)
